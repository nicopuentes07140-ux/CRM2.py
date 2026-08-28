
import os
from datetime import date
import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import sqlite3
 
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "produccion.db")
SEED_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "seed_marzo.csv")
 
MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}
MESES_INV = {v.lower(): k for k, v in MESES_ES.items()}
LINEAS = {"Samsung1": "SG(25,000 un)", "LG_TV": "LG(23,000 un)", "Neulsom LED": "NL_04 (60,000 un)"}
 
st.set_page_config(page_title="Seguimiento de Producción · Planta Sector tecnologico", layout="wide", page_icon="🏭")


def init_db():
    """Initialize the database used by the app."""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS produccion (
            fecha TEXT PRIMARY KEY,
            al01 INTEGER DEFAULT 0,
            al02 INTEGER DEFAULT 0,
            al04 INTEGER DEFAULT 0,
            num_cambios INTEGER DEFAULT 0,
            meta_dia INTEGER DEFAULT 0
        )
        """
    )
    conn.commit()
    conn.close()


def load_data():
    """Load all production records from the local SQLite database."""
    if not os.path.exists(DB_PATH):
        return pd.DataFrame()
    conn = sqlite3.connect(DB_PATH)
    try:
        df = pd.read_sql_query("SELECT * FROM produccion ORDER BY fecha", conn, parse_dates=["fecha"])
    except Exception:
        df = pd.DataFrame()
    conn.close()
    if df.empty:
        return df
    df["id"] = range(1, len(df) + 1)
    # compute derived fields
    df["produccion_total"] = df["al01"] + df["al02"] + df["al04"]
    df["mes_anio"] = pd.to_datetime(df["fecha"]).dt.to_period("M").astype(str)
    df["produccion_acumulada"] = df["produccion_total"].cumsum()
    df["meta_acumulada"] = df["meta_dia"].cumsum()
    df["cumplimiento"] = (df["produccion_acumulada"] / df["meta_acumulada"]).fillna(0)
    df["cumplimiento_dia"] = (df["produccion_total"] / df["meta_dia"]).replace([pd.NA, pd.NaT, float("inf")], 0).fillna(0)
    return df


def seed_marzo_demo(year: int):
    """Populate the database with demo production records for March."""
    if not os.path.exists(SEED_PATH):
        return None
    df_seed = pd.read_csv(SEED_PATH)
    # ensure fecha column exists
    if "fecha" not in df_seed.columns:
        return None
    conn = sqlite3.connect(DB_PATH)
    df_seed.to_sql("produccion", conn, if_exists="append", index=False)
    conn.close()
    return None


def reset_db():
    """Reset the database."""
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    init_db()
    return None


def fmt(x):
    try:
        return f"{int(x):,}"
    except Exception:
        return str(x)


def upsert_record(fecha, al01, al02, al04, num_cambios, meta_dia):
    """Insert or update a production record by fecha (ISO string)."""
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        "SELECT 1 FROM produccion WHERE fecha = ?",
        (fecha,),
    )
    exists = cur.fetchone() is not None
    if exists:
        cur.execute(
            """
            UPDATE produccion SET al01 = ?, al02 = ?, al04 = ?, num_cambios = ?, meta_dia = ? WHERE fecha = ?
            """,
            (al01, al02, al04, num_cambios, meta_dia, fecha),
        )
    else:
        cur.execute(
            """
            INSERT INTO produccion (fecha, al01, al02, al04, num_cambios, meta_dia) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (fecha, al01, al02, al04, num_cambios, meta_dia),
        )
    conn.commit()
    conn.close()

#  elaborar tabla csv
st.title("📁 Cargar CSV")
df = pd.read_csv("seed_marzo.csv")

def main():
    st.title("registro de produccion")
    st.header("Cargar datos de producción desde CSV")
    st.dataframe(df)

init_db()
main() 
# ---------------------------------------------------------------------------
# Sidebar / navegación
# ---------------------------------------------------------------------------
st.sidebar.title("🏭 Korea/Planta de Componentes tecnologicos")
st.sidebar.caption("Seguimiento de producción")
 
df_all = load_data()
 
if df_all.empty:
    st.sidebar.info("Todavía no hay datos cargados.")
    with st.sidebar.expander("Cargar datos de ejemplo (marzo)"):
        demo_year = st.number_input("Año", value=date.today().year, step=1, key="demo_year")
        if st.button("Cargar marzo de ejemplo"):
            seed_marzo_demo(int(demo_year))
            st.rerun()
 
page = st.sidebar.radio(
    "Navegación",
    ["📊 Dashboard", "➕ Registrar producción", "🗂️ Datos históricos", "📁 Importar Excel"],
)
 
with st.sidebar.expander("⚙️ Opciones avanzadas"):
    if st.button("🗑️ Borrar todos los datos"):
        reset_db()
        st.rerun()
 
st.sidebar.caption(f"Registros totales: {len(df_all)}")

# Página: Dashboard
# ---------------------------------------------------------------------------
if page == "📊 Dashboard":
    st.title("📊 Dashboard de producción")

    st.header("📁 Cargar archivo Excel y generar dashboard")
    archivo_excel = st.file_uploader("Sube tu archivo Excel (.xlsx)", type=["xlsx"])
    if archivo_excel is not None:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
        hoja_sel = st.selectbox("Selecciona la hoja", wb.sheetnames)
        df = pd.read_excel(archivo_excel, sheet_name=hoja_sel)
        st.success(f"Se cargaron {len(df)} filas de la hoja '{hoja_sel}'.")
        st.dataframe(df, use_container_width=True, hide_index=True)
        st.divider()
        st.subheader("⚙️ Configura el dashboard")
        columnas = list(df.columns)
        col_dia = st.selectbox("Columna del eje X (día)", columnas, index=0)
        col_lineas = st.multiselect("Columnas de producción por línea", [c for c in columnas if c != col_dia])
        col_meta = st.selectbox("Columna de meta del día (opcional)", ["(ninguna)"] + columnas)
        if col_lineas:
            df["produccion_total"] = df[col_lineas].sum(axis=1)
            st.subheader("Producción diaria por línea")
            df_melt = df.melt(id_vars=[col_dia], value_vars=col_lineas, var_name="línea", value_name="unidades")
            fig1 = px.bar(df_melt, x=col_dia, y="unidades", color="línea", barmode="stack")
            st.plotly_chart(fig1, use_container_width=True)
            st.subheader("Producción total por día")
            fig2 = px.line(df, x=col_dia, y="produccion_total", markers=True)
            st.plotly_chart(fig2, use_container_width=True)
            if col_meta != "(ninguna)":
                df["produccion_acumulada"] = df["produccion_total"].cumsum()
                df["meta_acumulada"] = df[col_meta].cumsum()
                df["cumplimiento"] = df["produccion_acumulada"] / df["meta_acumulada"] * 100
                st.subheader("Producción acumulada vs meta")
                fig3 = go.Figure()
                fig3.add_trace(go.Scatter(x=df[col_dia], y=df["produccion_acumulada"], name="Producción acumulada", mode="lines+markers"))
                fig3.add_trace(go.Scatter(x=df[col_dia], y=df["meta_acumulada"], name="Meta acumulada", mode="lines", line=dict(dash="dash")))
                st.plotly_chart(fig3, use_container_width=True)
                st.subheader("Cumplimiento diario (%)")
                df["cumplimiento_dia"] = df["produccion_total"] / df[col_meta] * 100
                fig4 = px.line(df, x=col_dia, y="cumplimiento_dia", markers=True)
                fig4.update_layout(yaxis_title="% cumplimiento")
                st.plotly_chart(fig4, use_container_width=True)

 #---------------------------------------------------------------------------------------------------------------------- 
    if df_all.empty:
        st.warning("Aún no hay datos. Usa 'Registrar producción' o 'Importar Excel' para cargar información.")
    else:
        meses_disponibles = sorted(df_all["mes_anio"].unique(), reverse=True)
        mes_sel = st.selectbox("Mes", meses_disponibles)
        df = df_all[df_all["mes_anio"] == mes_sel].copy()
 
        threshold = st.sidebar.slider("Umbral de alerta de cumplimiento", 50, 100, 80, help="% mínimo esperado por día") / 100
 
        ultima = df.iloc[-1]
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Producción acumulada", fmt(ultima["produccion_acumulada"]))
        c2.metric("Meta acumulada", fmt(ultima["meta_acumulada"]))
        c3.metric("Cumplimiento acumulado", f"{ultima['cumplimiento'] * 100:.1f}%")
        c4.metric("Promedio diario", fmt(df["produccion_total"].mean()))
 
        st.subheader("Producción diaria por línea")
        df_melt = df.melt(id_vars=["fecha"], value_vars=["al01", "al02", "al04"], var_name="linea", value_name="unidades")
        df_melt["linea"] = df_melt["linea"].map(LINEAS)
        fig1 = px.bar(df_melt, x="fecha", y="unidades", color="linea", barmode="stack")
        fig1.update_layout(legend_title_text="Línea", xaxis_title="", yaxis_title="Unidades")
        st.plotly_chart(fig1, use_container_width=True)
 
        col_a, col_b = st.columns(2)
        with col_a:
            st.subheader("Producción acumulada vs meta")
            fig2 = go.Figure()
            fig2.add_trace(go.Scatter(x=df["fecha"], y=df["produccion_acumulada"], name="Producción acumulada", mode="lines+markers"))
            fig2.add_trace(go.Scatter(x=df["fecha"], y=df["meta_acumulada"], name="Meta acumulada", mode="lines", line=dict(dash="dash")))
            fig2.update_layout(xaxis_title="", yaxis_title="Unidades")
            st.plotly_chart(fig2, use_container_width=True)
 
        with col_b:
            st.subheader("Cumplimiento diario (%)")
            fig3 = go.Figure()
            fig3.add_trace(go.Scatter(x=df["fecha"], y=df["cumplimiento_dia"] * 100, name="Cumplimiento del día", mode="lines+markers"))
            fig3.add_hline(y=threshold * 100, line_dash="dot", line_color="red", annotation_text=f"Umbral {int(threshold * 100)}%")
            fig3.add_hline(y=100, line_dash="dot", line_color="green", annotation_text="Meta 100%")
            fig3.update_layout(xaxis_title="", yaxis_title="%")
            st.plotly_chart(fig3, use_container_width=True)
 
        st.subheader("Resumen semanal")
        df["semana"] = pd.to_datetime(df["fecha"]).dt.isocalendar().week
        weekly = df.groupby("semana").agg(produccion=("produccion_total", "sum"), meta=("meta_dia", "sum")).reset_index()
        weekly["cumplimiento"] = (weekly["produccion"] / weekly["meta"] * 100).round(1)
        fig4 = px.bar(weekly, x="semana", y=["produccion", "meta"], barmode="group")
        fig4.update_layout(xaxis_title="Semana", yaxis_title="Unidades", legend_title_text="")
        st.plotly_chart(fig4, use_container_width=True)
 
        st.subheader("⚠️ Días bajo el umbral de cumplimiento")
        alertas = df[df["cumplimiento_dia"] < threshold][["fecha", "produccion_total", "meta_dia", "cumplimiento_dia"]].copy()
        if alertas.empty:
            st.success("Ningún día por debajo del umbral en este mes.")
        else:
            alertas["fecha"] = alertas["fecha"].dt.date
            alertas["cumplimiento_dia"] = (alertas["cumplimiento_dia"] * 100).round(1).astype(str) + "%"
            alertas.columns = ["Fecha", "Producción", "Meta día", "Cumplimiento"]
            st.dataframe(alertas, use_container_width=True, hide_index=True)

# ---------------------------------------------------------------------------
# Página: Registrar producción
# ---------------------------------------------------------------------------
elif page == "➕ Registrar producción":
    st.title("➕ Registrar producción diaria")
    st.caption("Si la fecha ya existe, el registro se actualiza (no se duplica).")
 
    with st.form("registro_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            fecha_in = st.date_input("Fecha", value=date.today())
            al01_in = st.number_input("AL-01 (unidades)", min_value=0, step=100)
            al02_in = st.number_input("AL-02 (unidades)", min_value=0, step=100)
        with col2:
            al04_in = st.number_input("AL-04 (unidades)", min_value=0, step=100)
            num_cambios_in = st.number_input("Número de cambios", min_value=0, step=1)
            meta_dia_in = st.number_input("Meta del día", min_value=0, step=500)
 
        submitted = st.form_submit_button("💾 Guardar registro")
        if submitted:
            upsert_record(fecha_in.isoformat(), al01_in, al02_in, al04_in, num_cambios_in, meta_dia_in)
            total = al01_in + al02_in + al04_in
            st.success(f"Registro guardado para {fecha_in.isoformat()} · Producción total del día: {fmt(total)} unidades")
            st.rerun()

#pagina:  Datos históricos
# ---------------------------------------------------------------------------
elif page == "🗂️ Datos históricos":
    st.title("🗂️ Datos históricos")
 
    if df_all.empty:
        st.info("No hay datos cargados todavía.")
    else:
        edit_df = df_all[["id", "fecha", "al01", "al02", "al04", "num_cambios", "meta_dia"]].copy()
        edit_df["fecha"] = edit_df["fecha"].dt.date
 
        st.caption("Puedes editar valores, añadir filas o borrarlas (selecciona la fila y presiona la tecla Supr). Luego guarda los cambios.")
        edited = st.data_editor(
            edit_df,
            num_rows="dynamic",
            use_container_width=True,
            hide_index=True,
            column_config={"id": st.column_config.NumberColumn("ID", disabled=True)},
        )
 
        col_save, _ = st.columns([1, 4])
        with col_save:
            if st.button("💾 Guardar cambios"):
                to_save = edited.copy()
                to_save["fecha"] = pd.to_datetime(to_save["fecha"], errors="coerce").dt.strftime("%Y-%m-%d")
                to_save = to_save.dropna(subset=["fecha"])
                replace_all(to_save)
                st.success("Cambios guardados.")
                st.rerun()
 
        st.divider()
        st.subheader("⬇️ Exportar")
        csv = df_all.drop(columns=["mes_anio"]).to_csv(index=False).encode("utf-8")
        st.download_button("Descargar CSV", csv, "produccion_historico.csv", "text/csv")

# Página: Importar Excel
# ---------------------------------------------------------------------------
elif page == "📁 Importar Excel":
    st.title("📁 Importar informe de Excel")
    st.caption(
        "Sube un archivo con el mismo formato de la hoja 'PROYECCIÓN GENERAL ...': "
        "columna 'Dia', producción por línea, número de cambios y meta día."
    )
    uploaded = st.file_uploader("Archivo Excel (.xlsx)", type=["xlsx"])
 
    if uploaded:
        wb = openpyxl.load_workbook(uploaded, data_only=True)
        hoja_sel = st.selectbox("Hoja con el detalle diario", wb.sheetnames)
        ws = wb[hoja_sel]
        rows = list(ws.iter_rows(min_col=2, values_only=True))
 
        header_idx = None
        for i, row in enumerate(rows):
            if row and isinstance(row[0], str) and row[0].strip().lower() == "dia":
                header_idx = i
                break
 
        if header_idx is None:
            st.error("No se encontró una fila de encabezado con 'Dia' en la columna B. Verifica el formato del archivo.")
        else:
            mes_guess = None
            for nombre, num in MESES_INV.items():
                if nombre in hoja_sel.lower():
                    mes_guess = num
                    break
 
            col1, col2 = st.columns(2)
            with col1:
                mes_sel = st.selectbox("Mes de estos datos", list(MESES_ES.values()), index=(mes_guess - 1) if mes_guess else 0)
            with col2:
                anio_sel = st.number_input("Año", value=date.today().year, step=1)
            mes_num = MESES_INV[mes_sel.lower()]
 
            registros = []
            for row in rows[header_idx + 1:]:
                if not row or row[0] is None or not isinstance(row[0], (int, float)):
                    continue
                dia = int(row[0])
                al01 = row[1] or 0
                al02 = row[2] or 0
                al04 = row[3] or 0
                num_cambios = row[5] if len(row) > 5 and row[5] is not None else 0
                meta_dia = row[7] if len(row) > 7 and row[7] is not None else 0
                try:
                    fecha = date(int(anio_sel), mes_num, dia).isoformat()
                except ValueError:
                    continue
                registros.append((fecha, al01, al02, al04, num_cambios, meta_dia))
 
            st.write(f"Se detectaron **{len(registros)}** días con datos.")
            preview = pd.DataFrame(registros, columns=["fecha", "al01", "al02", "al04", "num_cambios", "meta_dia"])
            st.dataframe(preview, use_container_width=True, hide_index=True)
 
            if registros and st.button("✅ Importar estos registros"):
                for r in registros:
                    upsert_record(*r)
                st.success(f"Se importaron {len(registros)} registros.")
                st.rerun()
 
 
